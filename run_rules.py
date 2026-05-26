"""
FAR Extraction Engine — run_rules.py  (v3)
==========================================
Applies extraction rules from extraction_rules.md to produce a
ground-truth-format Excel file from faculty annual report files.

Sources per faculty:
  1. F180Vita_<Initial>.<Last>.pdf  — FAR PDF
  2. <Last>.xlsx                    — Supplemental metrics (SDF sheet)
  3. <Last> CV.pdf                  — CV PDF
  4. Faculty Support Staff (PostDoc and GARs).xlsx

Usage:
  python run_rules.py                        # all known faculty
  python run_rules.py Narayanan              # single faculty
  python run_rules.py all --debug            # verbose

Dependencies:
  pip install pdfplumber openpyxl requests
"""

import sys, os, re, math, glob, argparse
from datetime import date
from collections import Counter
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────
REPORT_YEAR = 2024
Q4_START    = date(REPORT_YEAR, 10, 1)

RESEARCH_SHELL_TOKENS = {
    'RESEARCH', 'SEMINAR', 'INDEPENDENT', 'INTERNSHIP',
    'DIRECTED', 'PRACTICUM', 'SPECIAL', 'THESIS', 'DISSERTATION',
}
HONORS_TOKENS = {'HNR', 'HONORS'}

JOURNAL_HDR_KW = {
    'refereed journal', 'journal article', 'archival journal',
    'peer-reviewed journal', 'journal papers', 'journal publications',
}
CONF_HDR_KW = {
    'conference proceeding', 'conference paper', 'conference publication',
    'refereed conference', 'peer-reviewed conference',
}
JOURNAL_STATUS = {'published', 'accepted', 'in press', 'to appear'}

DEBUG = False

def dbg(msg):
    if DEBUG:
        print(f"  [DBG] {msg}")


# ── PDF helpers ────────────────────────────────────────────────────────────────

def pdf_full_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"
    return text

def pdf_all_tables(pdf_path):
    """Return list of (page_num, table) tuples from every page."""
    result = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            for tbl in (page.extract_tables() or []):
                result.append((i, tbl))
    return result


# ── Date / semester utils ──────────────────────────────────────────────────────

def semester_year(s):
    m = re.search(r'\b(\d{4})\b', str(s))
    return int(m.group(1)) if m else None

def date_contains_year(d_str, year):
    return str(year) in str(d_str)

def parse_date(d_str):
    if not d_str: return None
    # PDF cells often embed newlines: '2024-\n05-01' — strip all whitespace first
    d_str = re.sub(r'\s+', '', str(d_str))
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', d_str)
    if m:
        try: return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except: pass
    return None


# ── Table header-row finder ────────────────────────────────────────────────────

def _find_header_row(tbl, keywords, max_scan=4):
    """
    Return the row index (0-based) of the first row whose concatenated text
    contains ALL keywords (case-insensitive). Returns -1 if not found.
    Normalises embedded newlines so 'Student\\nName' matches 'student name'.
    """
    for i, row in enumerate(tbl[:max_scan]):
        rt = ' '.join(re.sub(r'\s+', ' ', str(c or '')).lower().strip() for c in row)
        if all(kw in rt for kw in keywords):
            return i
    return -1


def _cell(row, idx, default=''):
    """Safe cell access."""
    if idx < len(row):
        return str(row[idx] or '').strip()
    return default


# ── FAR: Teaching table ────────────────────────────────────────────────────────

def _extract_teaching(all_tables, report_year):
    """
    Returns list of dicts: {course_num, title, semester}
    Handles both primary tables (with 'Course Title' header) and the
    fact that the header is always in row 1 (row 0 = section title).
    """
    results = []
    for _, tbl in all_tables:
        if not tbl or len(tbl) < 2:
            continue
        h = _find_header_row(tbl, ['course title'])
        if h < 0:
            continue
        # Column indices from header row
        hrow = [str(c or '').lower().strip() for c in tbl[h]]
        # Find course-number col: header text 'course' (but not 'course title')
        course_col = 1  # default: col1 = "Course" (number)
        sem_col    = next((j for j, x in enumerate(hrow) if 'semester' in x), 10)
        for row in tbl[h + 1:]:
            title_raw = _cell(row, 0)
            course_raw = _cell(row, course_col)
            sem_raw   = _cell(row, sem_col)
            if not title_raw or not course_raw:
                continue
            # Skip header-repeat rows
            if 'course title' in title_raw.lower():
                continue
            m = re.search(r'\b(\d{3,4})\b', course_raw)
            if not m:
                continue
            yr = semester_year(sem_raw)
            if yr != report_year:
                continue
            results.append({
                'course_num': int(m.group(1)),
                'title':      re.sub(r'\s+', ' ', title_raw).upper(),
                'semester':   sem_raw,
            })
    dbg(f"Teaching rows: {len(results)}")
    return results


# ── FAR: Grad Advising table ───────────────────────────────────────────────────

def _extract_grad_advising(all_tables):
    """
    Returns list of dicts: {name, degree, major, grad_date, role}

    Two table formats appear:
      PRIMARY   : col0=Status, col1=StudentName, col5=Degree, col6=Major,
                  col7=GradDate, col13=Role
      CONTINUATION: col0=fused/empty, col1=Status, col2=StudentName, col6=Degree,
                  col7=Major, col8=GradDate, col14=Role
    """
    results = []
    seen_names = set()

    for _, tbl in all_tables:
        if not tbl or len(tbl) < 2:
            continue

        # Try to find header row
        h = _find_header_row(tbl, ['student name'])
        if h >= 0:
            # Primary format — derive col offsets from header
            hrow = [re.sub(r'\s+', ' ', str(c or '')).lower().strip() for c in tbl[h]]
            def col_idx(kw, default):
                for j, x in enumerate(hrow):
                    if kw in x: return j
                return default
            name_col    = col_idx('student name', 1)
            deg_col     = col_idx('degree', 5)
            maj_col     = col_idx('major', 6)
            grad_col    = col_idx('graduation', 7)
            role_col    = col_idx('role', 13)
            # End Semester column (used for activity detection)
            end_sem_col = col_idx('end semester', col_idx('end sem', -1))

            for row in tbl[h + 1:]:
                _parse_advising_row(row, name_col, deg_col, maj_col,
                                    grad_col, role_col, results, seen_names,
                                    end_sem_col=end_sem_col)
        else:
            # No explicit header — check if this looks like a grad continuation
            # by testing whether col1 starts with 'AS' or 'IS'
            is_continuation = any(
                re.match(r'^(AS|IS)\s*[-–]', str((row[1] if len(row) > 1 else '') or ''))
                for row in tbl[:3] if row
            )
            if not is_continuation:
                continue
            # Continuation format: col0=fused, col1=Status, col2=StudentName,
            # col6=Degree, col7=Major, col8=GradDate, col14=Role
            for row in tbl:
                _parse_advising_row(row, 2, 6, 7, 8, 14, results, seen_names)

    dbg(f"Grad advising rows: {len(results)}")
    return results


def _parse_advising_row(row, name_col, deg_col, maj_col, grad_col, role_col,
                        results, seen_names, end_sem_col=-1):
    name  = _cell(row, name_col)
    deg   = _cell(row, deg_col)
    maj   = _cell(row, maj_col)
    grad  = _cell(row, grad_col)
    role  = _cell(row, role_col)
    end_sem = _cell(row, end_sem_col) if end_sem_col >= 0 else ''

    # Normalise embedded newlines first
    name = re.sub(r'\s+', ' ', name).strip()

    # Strip program-code prefix: 'PHD-SC Allyson Larsen' → 'Allyson Larsen'
    # Matches patterns like PHD-SC, PHD-EN, MS-EN-THO, MENG-INTERDISC, etc.
    name = re.sub(r'^(PHD|MS|MEN|MENG)[-\w]*\s+', '', name, flags=re.IGNORECASE).strip()

    # Clean name: remove status prefix, inactive markers
    name = re.sub(r'^(AS|IS)\s*[-–]\s*(Active|Inactive)[,\s]*', '', name,
                  flags=re.IGNORECASE).strip()
    name = re.sub(r'\s*(Degree Candidate|Formal|Inactive)\s*$', '', name,
                  flags=re.IGNORECASE).strip()
    name = re.sub(r'^[,.\s]+', '', name).strip()

    # Reject: blank, too long, looks like a thesis title, contains digits
    if not name: return
    if len(name) > 50: return
    words = name.split()
    if len(words) > 5: return
    if re.search(r'[\d!?]', name): return
    if re.search(r'[.]{2,}', name): return
    # Reject pure lowercase garbage ('postdoctoralresearchers', etc.)
    if name == name.lower() and len(name) > 4: return

    key  = name.lower()
    if key in seen_names: return
    seen_names.add(key)

    results.append({
        'name':      name,
        'degree':    re.sub(r'\s+', ' ', deg).upper(),
        'major':     re.sub(r'\s+', ' ', maj).upper(),
        'grad_date': grad,
        'end_sem':   re.sub(r'\s+', ' ', end_sem).strip(),
        'role':      re.sub(r'\s+', ' ', role).strip(),
    })


# ── FAR: Grants table ─────────────────────────────────────────────────────────

def _extract_grants(all_tables, last_name):
    """
    Returns list of dicts: {title, role, start, end, status}

    Two formats:
      PRIMARY    : tbl[0] = column headers with 'Title' at col0
                   → col0=Title, col1=Collab, col3=Start, col4=End, col7=Status
      CONTINUATION: no header row; col0=fused/empty, col1=Title,
                   col2=Collab, col4=Start, col5=End, col8=Status
    """
    surname = last_name.upper() if last_name else ''
    results = []
    seen_titles = set()

    for _, tbl in all_tables:
        if not tbl or len(tbl) < 1:
            continue

        # Detect format by looking for a header row
        h = _find_header_row(tbl, ['title', 'collaborat'])
        if h >= 0 and str(tbl[h][0] or '').strip().lower() == 'title':
            # PRIMARY format: header at row h, data from h+1
            for row in tbl[h + 1:]:
                _parse_grant_row_primary(row, surname, results, seen_titles)
        else:
            # Check if this looks like a grant continuation table:
            # rows should have date patterns — normalize whitespace first so
            # '2024-\n05-01' becomes '2024-05-01' before matching.
            looks_like_grants = any(
                re.search(r'20\d\d-\d\d-\d\d',
                          re.sub(r'\s+', '', ' '.join(str(c or '') for c in row)))
                for row in tbl[:3]
            )
            if not looks_like_grants:
                continue
            for row in tbl:
                _parse_grant_row_continuation(row, surname, results, seen_titles)

    dbg(f"Grant rows: {len(results)}")
    return results


def _parse_grant_row_primary(row, surname, results, seen_titles):
    """Primary format: Title=col0, Collab=col1, Start=col3, End=col4, Status=col7"""
    if len(row) < 8: return
    title  = _cell(row, 0)
    collab = _cell(row, 1)
    start  = _cell(row, 3)
    end    = _cell(row, 4)
    status = _cell(row, 7)
    _add_grant(title, collab, start, end, status, surname, results, seen_titles)


def _parse_grant_row_continuation(row, surname, results, seen_titles):
    """Continuation format: fused col0, Title=col1, Collab=col2, Start=col4, End=col5, Status=col8"""
    if len(row) < 9: return
    # Reject rows that are clearly header rows (col1 == 'Title')
    if str(row[1] or '').strip().lower() == 'title':
        return
    # Skip fused rows with >3 dates (multiple grants merged into one cell)
    all_text = ' '.join(str(c or '') for c in row)
    if len(re.findall(r'\d{4}-\d{2}-\d{2}', all_text)) > 3:
        return
    title  = _cell(row, 1)
    collab = _cell(row, 2)
    start  = _cell(row, 4)
    end    = _cell(row, 5)
    status = _cell(row, 8)
    _add_grant(title, collab, start, end, status, surname, results, seen_titles)


def _add_grant(title, collab, start, end, status, surname, results, seen_titles):
    title  = re.sub(r'\s+', ' ', title).strip()
    status = re.sub(r'\s+', ' ', status).strip()
    if not title or len(title) < 2: return
    # Reject header-repeat rows
    if title.lower() in ('title', 'sponsored research'): return

    role = _find_role_in_collaborators(collab, surname)
    # Dedup key includes start date so different-year renewals of the same
    # project are treated as distinct rows (e.g. 2021 vs 2024 awards).
    start_key = re.sub(r'\s+', '', start)[:10]   # first 10 non-ws chars of start
    canon = _canonical_title(title)
    key = f"{canon}|{start_key}"
    if key in seen_titles: return
    seen_titles.add(key)

    results.append({
        'title':  title,
        'role':   role,
        'start':  start,
        'end':    end,
        'status': status,
    })


def _find_role_in_collaborators(collab_text, last_name):
    if not last_name or not collab_text:
        return 'Unknown'
    text = re.sub(r'\s+', ' ', collab_text)
    # Find last name followed by the immediate role token (single word only)
    pattern = re.compile(
        r'\b' + re.escape(last_name) + r'\b\s*([A-Za-z\-]+)',
        re.IGNORECASE
    )
    m = pattern.search(text)
    if m:
        tok = m.group(1).strip().lower()
        # Match only well-known role keywords
        if re.search(r'^pi$|^principal\s*invest', tok): return 'PI'
        if re.search(r'^co.?pi$|^co.?invest|^coinvest', tok): return 'CoPI'
        if re.search(r'^other$|^collaborat', tok): return 'Other'
        # Unknown single token (could be first name like 'Nicholas') → Other
        return 'Other'
    if last_name.upper() in text.upper():
        return 'Unknown-present'
    return 'Unknown-absent'


def _canonical_title(title):
    t = re.sub(r'^[A-Z][\w\s]+/', '', title)   # subaward prefix
    t = re.sub(r'\s*[\(\[]\s*FY\s*\d+.*?[\)\]]', '', t, flags=re.IGNORECASE)
    return t.strip().lower()


# ── FAR publications (text-based) ─────────────────────────────────────────────

def _parse_far_pubs(text, data):
    """
    Parse FAR publication text sections.
    Publications are numbered entries (e.g. "1. Author, ...").
    Only lines that START with a number followed by a period are counted as
    new entries; continuation lines (which may contain a year from the venue
    name) are accumulated onto the current entry and not double-counted.
    """
    lines      = text.split('\n')
    in_journal = False
    in_conf    = False
    year_pat   = re.compile(r'\b(20\d{2})\b')
    entry_re   = re.compile(r'^\d+\.')   # lines that start a numbered entry

    # Buffer for the current multi-line entry
    cur_entry_lines = []
    cur_section     = None   # 'journal' | 'conf' | None

    def _flush():
        if not cur_entry_lines or cur_section is None:
            return
        full = ' '.join(cur_entry_lines)
        m = year_pat.search(full)
        if m:
            yr = int(m.group(1))
            if cur_section == 'journal':
                data['journal_pubs'].append({'year': yr, 'text': full[:200]})
            elif cur_section == 'conf':
                data['conf_pubs'].append({'year': yr, 'text': full[:200]})

    for line in lines:
        ll      = line.lower()
        stripped = line.strip()

        # Section-header detection
        if any(h in ll for h in ('journal article', 'journal paper', 'refereed journal',
                                  'archival journal', 'peer-reviewed journal', 'journal publications')):
            _flush(); cur_entry_lines = []; cur_section = None
            in_journal = True; in_conf = False; continue
        if any(h in ll for h in ('conference proceeding', 'conference paper',
                                  'refereed conference', 'peer-reviewed conference',
                                  'conference publication')):
            _flush(); cur_entry_lines = []; cur_section = None
            in_conf = True; in_journal = False; continue
        if re.match(r'^(grants|graduate advising|service|goals|professional)', ll):
            _flush(); cur_entry_lines = []; cur_section = None
            in_journal = False; in_conf = False

        if not stripped:
            continue

        if entry_re.match(stripped):
            # New numbered entry — flush previous, start fresh
            _flush()
            cur_entry_lines = [stripped]
            cur_section = 'journal' if in_journal else ('conf' if in_conf else None)
        elif cur_entry_lines and cur_section is not None:
            # Continuation of the current entry
            cur_entry_lines.append(stripped)

    _flush()  # don't forget the last entry


# ── FAR header parser ──────────────────────────────────────────────────────────

def _parse_far_header(lines):
    """Extract name, title, report_year from the first ~25 lines of FAR text."""
    name = first_name = last_name = ''
    title = 'Professor'
    report_year = REPORT_YEAR

    for line in lines[:25]:
        ls = line.strip()
        if not ls: continue
        # Year detection
        if re.search(r'\b(Spring|Fall|Summer)\b', ls) and re.search(r'\d{4}', ls):
            yrs = re.findall(r'\d{4}', ls)
            if yrs: report_year = int(yrs[-1])
        # Title detection
        if re.search(r'\bProfessor\b', ls):
            if 'Associate' in ls:
                title = 'Associate Professor'
            elif 'Assistant' in ls:
                title = 'Assistant Professor'
            else:
                title = 'Professor'
        # Name: first non-empty line without dates/email/form keywords
        if not name:
            if (not re.search(r'Spring|Fall|Summer|\d{4}|@|Faculty|Form|Department|Title|Rank', ls)
                    and re.match(r'[A-Z]', ls)):
                parts = ls.split()
                if 2 <= len(parts) <= 6:
                    first_name = parts[0]
                    last_name  = parts[-1]
                    name = ls

    return name, first_name, last_name, title, report_year


# ── Main FAR parser ────────────────────────────────────────────────────────────

def parse_far(pdf_path):
    data = {
        'name': '', 'first_name': '', 'last_name': '', 'title': 'Professor',
        'report_year': REPORT_YEAR,
        'teaching':      [],
        'grants':        [],
        'grad_advising': [],
        'journal_pubs':  [],
        'conf_pubs':     [],
    }
    full_text = pdf_full_text(pdf_path)
    lines     = full_text.split('\n')
    all_tables = pdf_all_tables(pdf_path)

    name, fn, ln, title, yr = _parse_far_header(lines)
    data.update(name=name, first_name=fn, last_name=ln, title=title, report_year=yr)

    data['teaching']      = _extract_teaching(all_tables, data['report_year'])
    data['grad_advising'] = _extract_grad_advising(all_tables)
    data['grants']        = _extract_grants(all_tables, ln)
    _parse_far_pubs(full_text, data)

    return data


# ── CV parser ─────────────────────────────────────────────────────────────────

def parse_cv(pdf_path):
    result = {
        'current_advisees':  set(),
        'journal_pubs_2024': 0,
        'conf_pubs_2024':    0,
    }
    if not pdf_path or not os.path.exists(pdf_path):
        return result

    full_text = pdf_full_text(pdf_path)
    lines     = full_text.split('\n')

    _parse_cv_advisees(lines, result)
    _parse_cv_pubs(lines, full_text, result)

    return result


# ── CV advisees ────────────────────────────────────────────────────────────────

def _parse_cv_advisees(lines, data):
    full_text = '\n'.join(lines)

    # ── Pattern 1: Numbered "Current Students: N. Name (year-present)" list (Qian-style) ──
    cs_match = re.search(
        r'[–\-]\s*Current Students:\s*(.*?)(?=[–\-]\s*Alumni:|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if cs_match:
        block = cs_match.group(1)
        # Extract "Name(year-present)" or "Name (year - present)" entries
        # Names can be fused like "ZiyuXiang" or spaced like "Seyednami Niyakan"
        for m in re.finditer(
            r'\d+\.\s*([A-Z][A-Za-z\-]+(?:\s+[A-Z][A-Za-z\-]+)*)\s*\([\d\s]+[–\-]\s*present',
            block, re.IGNORECASE
        ):
            full_name = m.group(1).strip()
            # Handle fused CamelCase names: split on internal capitals
            parts = re.findall(r'[A-Z][a-z\-]*', full_name)
            if parts:
                surname = parts[-1].lower()
                if len(surname) > 2:
                    data['current_advisees'].add(surname)
        # Also check Master's current students section
        ms_match = re.search(
            r"Master['']s Students.*?[–\-]\s*Current Students:\s*(.*?)(?=[–\-]\s*Alumni:|\Z)",
            full_text, re.IGNORECASE | re.DOTALL
        )
        if ms_match:
            ms_block = ms_match.group(1)
            for m in re.finditer(
                r'\d+\.\s*([A-Z][A-Za-z\-]+(?:\s+[A-Z][A-Za-z\-]+)*)\s*[,\(]',
                ms_block, re.IGNORECASE
            ):
                full_name = m.group(1).strip()
                parts = re.findall(r'[A-Z][a-z\-]*', full_name)
                if parts:
                    surname = parts[-1].lower()
                    if len(surname) > 2:
                        data['current_advisees'].add(surname)

    # ── Pattern 2: Bullet-list "Current Advisees" sections (standard style) ──
    CURRENT_SECTION = re.compile(
        r'current\s+(phd|grad|student|advis)|'
        r'student\s+advis.*(current|active)|'
        r'phd\s+students.*current|'
        r'active\s+phd|'
        r'theses?\s+(in\s+progress|advised.*in)',
        re.IGNORECASE
    )
    # Pattern 3: "GRADUATE STUDENT COMMITTEE CHAIR" — extract only lines with "expected"/"ongoing"
    CHAIR_SECTION = re.compile(r'graduate\s+student\s+committee\s+chair', re.IGNORECASE)

    in_current = False
    in_chair   = False
    for line in lines:
        stripped = line.strip()
        ll       = stripped.lower()

        # Section starts
        if CURRENT_SECTION.search(line) and not cs_match:
            in_current = True; in_chair = False; continue
        if CHAIR_SECTION.search(line):
            in_chair = True; in_current = False; continue

        # Section ends
        if in_current or in_chair:
            if (re.match(r'^[A-Z][A-Z\s]{5,}$', stripped)
                    and len(stripped) > 8 and not re.search(r'\d', stripped)):
                in_current = False; in_chair = False; continue
            if re.match(r'^(Graduated|Alumni|Former|Past|Education|Research|Teaching|Service|Honors)',
                        stripped, re.IGNORECASE):
                in_current = False; in_chair = False; continue

        # Extract from standard current-advisee sections
        if in_current:
            nm2 = re.match(r'^[\s•\-\*]*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3})[,–\(]', stripped)
            if nm2:
                full_name = nm2.group(1).strip()
                last = full_name.split()[-1]
                if len(last) > 2:
                    data['current_advisees'].add(last.lower())

        # Extract from CHAIR section: ONLY active students (lines with "expected"/"ongoing")
        if in_chair and ('expected' in ll or 'ongoing' in ll):
            nm2 = re.match(r'^[\s•\-\*]*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3})[,–\(]', stripped)
            if nm2:
                full_name = nm2.group(1).strip()
                last = full_name.split()[-1]
                if len(last) > 2:
                    data['current_advisees'].add(last.lower())

    # ── Pattern 4: Asterisk-legend convention (Palermo: "* Student or post-doc in Prof. X's group") ──
    has_asterisk_legend = bool(re.search(
        r'\*\s*[=:]?\s*(student|post.?doc)\s+(or\s+post.?doc\s+in|in\s+Prof)',
        full_text, re.IGNORECASE
    ))
    if has_asterisk_legend:
        # Alumni = graduated BEFORE report year (strictly <).
        # Students who graduated IN the report year still count as current
        # since they were active during that year.
        alumni_surnames = _cv_alumni_surnames(full_text, max_year=REPORT_YEAR - 1)

        # Find starred names in REPORT_YEAR pub ENTRIES (entry-level, not line-level).
        # Build a pub-section text from journal + conference sections, then split into entries.
        starred_names = set()
        year_re = re.compile(r'\b' + str(REPORT_YEAR) + r'\b')
        entry_pat = re.compile(r'^\d+\.\s', re.MULTILINE)

        # Collect the journal and conference section text
        pub_lines = []
        in_pub = False
        for line in lines:
            ll2 = line.lower()
            if any(kw in ll2 for kw in JOURNAL_HDR_KW) or any(kw in ll2 for kw in CONF_HDR_KW):
                in_pub = True; continue
            if re.match(r'^(thes|grant|service|teaching|education|invited|award|honor|patent)',
                        ll2.strip()):
                in_pub = False
            if in_pub:
                pub_lines.append(line)
        pub_text = '\n'.join(pub_lines)

        # Split into individual entries
        entry_spans = [(m.start(), m.group(0)) for m in entry_pat.finditer(pub_text)]
        for idx, (start, _) in enumerate(entry_spans):
            end = entry_spans[idx + 1][0] if idx + 1 < len(entry_spans) else len(pub_text)
            entry = pub_text[start:end]
            if not year_re.search(entry):
                # Accept "accepted" / "in press" entries as current year
                el = entry.lower()
                if not any(s in el for s in ('accepted', 'in press', 'to appear')):
                    continue
            # Extract single-starred (not double-starred) names
            for m in re.finditer(r'([A-Z][a-z]{1,})\*(?!\*)', entry):
                starred_names.add(m.group(1).lower())

        # Add starred names that are NOT alumni
        for nm in starred_names:
            if nm not in alumni_surnames:
                data['current_advisees'].add(nm)
        dbg(f"CV asterisk advisees (non-alumni): {starred_names - alumni_surnames}")
        dbg(f"CV alumni excluded: {alumni_surnames & starred_names}")

    dbg(f"CV advisees total: {data['current_advisees']}")


# ── CV publications ────────────────────────────────────────────────────────────

def _parse_cv_pubs(lines, full_text, data):
    """
    Section-based publication counting.
    Identifies journal and conference sections by header keywords, then
    counts numbered entries whose year matches report_year.
    """
    # Build section map: list of (line_idx, section_type)
    section_map = []
    for i, line in enumerate(lines):
        ll = line.lower()
        if any(kw in ll for kw in JOURNAL_HDR_KW):
            section_map.append((i, 'journal'))
        elif any(kw in ll for kw in CONF_HDR_KW):
            section_map.append((i, 'conference'))
        elif re.match(r'^(INVITED TALKS?|TALKS?|PRESENTATIONS?|BOOK CHAPTERS?|'
                      r'THESES?|GRANTS?|PATENTS?|EDITORIAL|SERVICE|TEACHING|HONORS)',
                      line.strip(), re.IGNORECASE):
            section_map.append((i, 'other'))

    if not section_map:
        dbg("CV: no section map found; skipping pub counting")
        return

    # Determine which section each line belongs to
    def section_at(line_idx):
        cur = 'other'
        for idx, stype in section_map:
            if idx <= line_idx:
                cur = stype
            else:
                break
        return cur

    # Pick the entry-splitting pattern (most splits in journal+conf sections)
    candidates = [
        re.compile(r'^(\d+)\.\s', re.MULTILINE),
        re.compile(r'^\[(\d+)\]\s', re.MULTILINE),
        re.compile(r'^\[([A-Z]\d+)\]\s', re.MULTILINE),
        re.compile(r'^(\d+)\s+[A-Z]', re.MULTILINE),    # Narayanan style: "1 Author..."
    ]
    best_pat = None
    best_count = 0
    for pat in candidates:
        cnt = len(pat.findall(full_text))
        if cnt > best_count:
            best_count = cnt
            best_pat   = pat
    if not best_pat or best_count < 2:
        dbg("CV: could not find entry pattern")
        return
    dbg(f"CV split pattern: {best_pat.pattern!r} ({best_count} entries)")

    # Find all entry start positions and their line indices
    entry_starts = [(m.start(), full_text[:m.start()].count('\n')) for m in best_pat.finditer(full_text)]

    if not entry_starts:
        return

    year_pat = re.compile(r'\b(20\d\d)\b')

    # Extract the text of each entry (from its start to the next entry's start)
    for i, (char_pos, line_idx) in enumerate(entry_starts):
        next_pos = entry_starts[i + 1][0] if i + 1 < len(entry_starts) else len(full_text)
        chunk = full_text[char_pos:next_pos].strip()

        # Determine section
        sec = section_at(line_idx)
        if sec == 'other':
            # Try keyword override — only for entries that look like authored papers.
            # An authored paper starts with an author name (e.g. "S. Palermo," or
            # "Smith, J.") immediately after the entry number.
            # Event listings start with an org/event name (IEEE, ACM, a year, or a
            # capitalized institution name) — skip those.
            cl = chunk.lower()
            # Extract text after the leading "N." or "[N]" entry marker
            body_m = re.match(r'^\[?\d+[\.\]]\s*(.*)', chunk, re.DOTALL)
            body_start = body_m.group(1)[:60] if body_m else chunk[:60]
            looks_like_paper = bool(re.match(
                r'[A-Z][a-z\-]+[\*,\s]|[A-Z]\.',   # author: "Smith," or "S." or "Kim*"
                body_start
            )) and not re.match(
                r'(IEEE|ACM|20\d{2}|[A-Z][A-Z\s]{4,})',  # org/year/all-caps name
                body_start
            )
            if looks_like_paper:
                if any(kw in cl for kw in JOURNAL_HDR_KW) or any(
                        w in cl for w in ('ieee transactions', 'ieee journal', 'acm transactions',
                                          'journal of ', 'letters on', 'transactions on')):
                    sec = 'journal'
                elif any(kw in cl for kw in CONF_HDR_KW) or any(
                        w in cl for w in ('proceedings', 'workshop', 'symposium',
                                          'conference', 'icml', 'neurips', 'iclr')):
                    sec = 'conference'

        if sec not in ('journal', 'conference'):
            continue

        # Reject dissertation / thesis entries
        cl = chunk.lower()
        if re.search(r'\b(dissertation|thesis|ph\.?d\.?\s+thesis|m\.s\.?\s+thesis)\b', cl):
            continue

        # Reject invited talks / presentations (not peer-reviewed conference papers)
        if re.search(r'\b(invited talk|keynote|plenary talk|talk at|tutorial|'
                     r'forum talk|workshop talk|conference talk|panel)\b', cl):
            continue
        # Also reject entries that look like conference/event announcements
        # rather than authored papers (no comma-separated author list before the title)
        if sec == 'conference' and re.search(r'\bTalk\b', chunk):
            continue

        # Reject grant listings (contain dollar amounts or "PI:" patterns)
        if re.search(r'\$[\d,]+|\bPI:\b|\bCo-PI:\b', chunk):
            continue

        # Find the year of this entry
        year_m = year_pat.search(chunk)
        if year_m:
            yr = int(year_m.group(1))
            if yr != REPORT_YEAR:
                # Allow entries published in REPORT_YEAR+1 when accepted/posted in
                # REPORT_YEAR — indicated by REPORT_YEAR appearing in the DOI path
                # (e.g. "doi.org/.../compag.2024.109589") or "accepted" status.
                if yr == REPORT_YEAR + 1:
                    cl2 = chunk.lower()
                    doi_accepted = bool(re.search(
                        r'doi\.org.*[./]' + str(REPORT_YEAR) + r'[./]', cl2
                    ))
                    # "accepted/in press" counts only if it is NOT "accepted in {next_year}"
                    # (which means the paper is for a future conference/journal, not a
                    # 2024 publication that happens to be formally released in 2025).
                    next_yr = str(REPORT_YEAR + 1)
                    accepted_current = (
                        any(s in cl2 for s in ('accepted', 'in press', 'to appear'))
                        and not re.search(
                            r'\b(?:accepted|appearing|to\s+appear)\s+(?:in\s+)?' + next_yr,
                            cl2)
                    )
                    if not doi_accepted and not accepted_current:
                        continue
                    yr = REPORT_YEAR
                else:
                    continue
        else:
            # No explicit year — count only if "accepted" / "in press" status present
            cl2 = chunk.lower()
            if not any(s in cl2 for s in ('accepted', 'in press', 'to appear')):
                continue
            yr = REPORT_YEAR  # assume current year for accepted/in-press entries

        if sec == 'journal':
            data['journal_pubs_2024'] += 1
            dbg(f"  CV journal 2024: {chunk[:80]}")
        elif sec == 'conference':
            data['conf_pubs_2024'] += 1
            dbg(f"  CV conf 2024: {chunk[:80]}")


# ── Supplemental xlsx ─────────────────────────────────────────────────────────

def read_supplemental(xlsx_path):
    result = {'ms_grad': 0, 'phd_grad': 0, 'journal_raw': 0, 'conf_total': 0}
    if not xlsx_path or not os.path.exists(xlsx_path):
        return result
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    conf_buckets = []
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        label = str(row[0] or '').lower().strip()
        val   = row[2] if len(row) > 2 else None
        if 'ms/men' in label or ('ms' in label and 'graduated' in label):
            result['ms_grad'] = _parse_xlsx_number(val)
        elif 'phd' in label and 'graduated' in label:
            result['phd_grad'] = _parse_xlsx_number(val)
        elif ('journal' in label and ('publication' in label or 'paper' in label)):
            result['journal_raw'] = _parse_xlsx_number(val)
        elif 'conference' in label and any(x in label for x in ('<30', '30-60', '>60', '60%', '30%')):
            conf_buckets.append(_parse_xlsx_number(val))
        elif 'conference' in label and 'submission' in label:
            conf_buckets.append(_parse_xlsx_number(val))
    result['conf_total'] = sum(conf_buckets)
    dbg(f"Supplemental: {result}")
    return result


def _parse_xlsx_number(val):
    if val is None: return 0
    s = str(val).strip()
    if not s or s.lower() == 'none': return 0
    m = re.match(r'(\d+(?:\.\d+)?)', s)
    if m: return math.floor(float(m.group(1)))
    return 0


# ── Support staff xlsx ─────────────────────────────────────────────────────────

def read_support_staff(xlsx_path, last_name, first_name=''):
    if not xlsx_path or not os.path.exists(xlsx_path):
        return 0, 0
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None: continue
        ln = str(row[0]).strip().upper()
        fn = str(row[1]).strip().upper() if len(row) > 1 else ''
        if ln == last_name.upper() or (first_name and fn.startswith(first_name[0].upper())):
            postdocs = int(row[3]) if len(row) > 3 and row[3] else 0
            gars     = int(row[4]) if len(row) > 4 and row[4] else 0
            return postdocs, gars
    return 0, 0


# ── Course-type predicates ─────────────────────────────────────────────────────

def is_research_shell(course):
    title = course['title'].upper()
    return any(tok in title for tok in RESEARCH_SHELL_TOKENS)

def is_honors_capstone(course):
    title = course['title'].upper()
    return any(h in title for h in HONORS_TOKENS) and not is_research_shell(course)


# ── Dept code helper ──────────────────────────────────────────────────────────

def faculty_home_dept(grad_advising):
    codes = []
    for r in grad_advising:
        deg = r.get('degree', '').upper()
        maj = r.get('major', '').strip().upper()
        if re.match(r'^(PHD|MS|MEN)', deg) and re.match(r'^[A-Z]{2,6}$', maj):
            codes.append(maj)
    if not codes:
        return 'ELEN'
    return Counter(codes).most_common(1)[0][0]


# ── Rules 1–8 ─────────────────────────────────────────────────────────────────

def _pure_capstone_numbers(teaching):
    """
    A course number is a 'pure' honors capstone (counts as Grad) only if it
    appears exclusively with HNR/HONORS in its title and never without.
    If both 'ECEN 325 HNR-ELECTRONICS' and 'ECEN 325 ELECTRONICS' appear,
    the 325 HNR section is just an honors section of a regular UG course.
    """
    hnr_nums  = {c['course_num'] for c in teaching
                 if any(h in c['title'].upper() for h in HONORS_TOKENS)
                 and not is_research_shell(c)}
    plain_nums = {c['course_num'] for c in teaching
                  if not any(h in c['title'].upper() for h in HONORS_TOKENS)
                  and not is_research_shell(c)}
    return hnr_nums - plain_nums   # only pure capstones (no non-HNR sibling)


def rule1_ug_courses(far):
    pure_cap = _pure_capstone_numbers(far['teaching'])
    seen = set()
    for c in far['teaching']:
        n = c['course_num']
        if n < 500 and not is_research_shell(c) and n not in pure_cap:
            seen.add(n)
    dbg(f"UG courses: {seen}")
    return len(seen)


def rule2_grad_courses(far):
    pure_cap = _pure_capstone_numbers(far['teaching'])
    seen = set()
    for c in far['teaching']:
        n = c['course_num']
        if not is_research_shell(c) and (n >= 500 or n in pure_cap):
            seen.add(n)
    dbg(f"Grad courses: {seen}")
    return len(seen)


def rule3_phd_graduated(far, xlsx, cv_text=''):
    ga = far['grad_advising']
    role_filled = sum(1 for r in ga if r['role'].strip())
    use_fallback = len(ga) > 0 and role_filled / len(ga) < 0.25

    if use_fallback:
        dbg("PhD: xlsx fallback (sparse FAR roles)")
        return xlsx.get('phd_grad', 0)

    count = 0
    for r in ga:
        deg  = r['degree'].upper()
        role = r['role'].strip()
        gd   = r['grad_date']
        if re.search(r'\bchair\b', role, re.IGNORECASE) and deg.startswith('PHD'):
            if date_contains_year(gd, REPORT_YEAR):
                name = r['name']
                if cv_text and not _name_in_cv(name, cv_text):
                    dbg(f"PhD: skipping {name} (not in CV)")
                    continue
                count += 1
                dbg(f"PhD: {name} / {deg} / {gd}")

    # Xlsx fallback: if FAR shows 0 PhD chair graduates but xlsx self-report > 0,
    # the FAR likely mislabelled the role (e.g. Palermo lists PhD students as 'Member')
    if count == 0 and xlsx.get('phd_grad', 0) > 0:
        dbg(f"PhD: FAR=0 but xlsx={xlsx['phd_grad']} → using xlsx fallback")
        return xlsx['phd_grad']

    return count


def rule4_ms_graduated(far, xlsx, cv_text=''):
    ga = far['grad_advising']
    role_filled = sum(1 for r in ga if r['role'].strip())
    use_fallback = len(ga) > 0 and role_filled / len(ga) < 0.25

    if use_fallback:
        dbg("MS: xlsx fallback (sparse FAR roles)")
        return xlsx.get('ms_grad', 0)

    dept = faculty_home_dept(ga)
    ms_pat = re.compile(r'^(MS|MEN)\s+' + re.escape(dept) + r'$', re.IGNORECASE)
    count = 0
    for r in ga:
        deg  = r['degree'].upper()
        maj  = r['major'].upper()
        combined = f"{deg} {maj}".strip()
        role = r['role'].strip()
        gd   = r['grad_date']
        if re.search(r'\bchair\b', role, re.IGNORECASE) and ms_pat.match(combined):
            if date_contains_year(gd, REPORT_YEAR):
                name = r['name']
                if cv_text and not _name_in_cv(name, cv_text):
                    dbg(f"MS: skipping {name} (not in CV)")
                    continue
                count += 1
                dbg(f"MS: {name} / {combined} / {gd}")
    return count


def _name_in_cv(full_name, cv_text):
    parts = full_name.split()
    if not parts: return True
    surname = parts[-1]
    return bool(re.search(r'\b' + re.escape(surname) + r'\b', cv_text, re.IGNORECASE))


def _parse_cv_gifts(cv_text):
    """
    Count industry/foundation 'gifts' listed in the CV FUNDING section under a
    'Gifts' sub-heading that started in REPORT_YEAR.

    CV format (Palermo-style):
        Gifts
        Org, "Title," PI: Name, M/YYYY, $Amount
        ...

    Returns the count of gifts whose date field contains REPORT_YEAR.
    """
    if not cv_text:
        return 0
    # Find the Gifts subsection
    m = re.search(r'\nGifts\s*\n', cv_text, re.IGNORECASE)
    if not m:
        return 0
    # Extract text from the Gifts heading until the next all-caps heading or end
    rest = cv_text[m.end():]
    section_end = re.search(r'\n[A-Z][A-Z\s]{4,}\n', rest)
    gifts_text = rest[:section_end.start()] if section_end else rest

    count = 0
    # Each gift line mentions a year somewhere (as "M/YYYY" or ", YYYY,")
    year_str = str(REPORT_YEAR)
    for line in gifts_text.split('\n'):
        line = line.strip()
        if not line:
            continue
        # Must be a substantive gift line (contains a dollar sign or org name)
        if '$' not in line and not re.search(r'\b[A-Z][a-z]', line):
            continue
        if year_str in line:
            count += 1
            dbg(f"CV Gift {REPORT_YEAR}: {line[:80]}")
    return count


def _cross_ref_grants(faculty_last_name, all_far_data):
    """
    Cross-reference other faculty FARs to find qualifying grants where
    `faculty_last_name` is listed as a PI/CoPI collaborator but the grant
    does NOT appear as 'Funded-In-Progress' in their own FAR.

    This handles the case where one faculty's FAR has stale status
    (e.g., "In Preparation") while the grant is confirmed funded in a
    co-investigator's FAR.

    Returns: count of additional cross-reference grants.
    """
    year_start = date(REPORT_YEAR, 1, 1)
    year_end   = date(REPORT_YEAR, 12, 31)
    added = set()

    for other_last, (other_far, other_far_text) in all_far_data.items():
        if other_last.lower() == faculty_last_name.lower():
            continue
        for g in other_far.get('grants', []):
            status = g['status'].lower()
            if 'funded' not in status or 'progress' not in status:
                continue
            role = g['role']
            if role not in ('PI', 'CoPI'):
                continue
            start = parse_date(g['start'])
            if not start or not (year_start <= start <= year_end):
                continue
            end = parse_date(g['end'])
            if end is not None and end < Q4_START:
                continue
            # Check if the faculty appears as collaborator in the raw FAR text
            # near this grant title.
            # In pdfplumber table extraction the title words are interleaved with
            # other column values (PI name, agency, dates…), so we search using
            # just the first distinctive word/phrase of the title instead of a
            # long substring.
            title_full = g['title'].strip()
            if not title_full or len(title_full) < 5:
                continue
            # Use first word that is ≥5 chars as the search anchor
            anchor = ''
            for word in re.split(r'\W+', title_full):
                if len(word) >= 5:
                    anchor = word
                    break
            if not anchor:
                anchor = title_full[:15]
            title_idx = other_far_text.lower().find(anchor.lower())
            if title_idx < 0:
                continue
            window = other_far_text[title_idx: title_idx + 800]
            if faculty_last_name.lower() in window.lower():
                canon = _canonical_title(g['title'])
                if canon not in added:
                    added.add(canon)
                    dbg(f"Cross-ref grant for {faculty_last_name}: "
                        f"{g['title'][:55]} (from {other_last})")
    return len(added)


def rule5_grants(far, cv_text='', faculty_last_name='', all_far_data=None):
    """
    Count distinct PI/CoPI grants that are Funded-In-Progress,
    started within the report year, AND still active through Q4.

    Also counts:
      - CV 'Gifts' that started in REPORT_YEAR (industry gifts not in FAR table)
      - Cross-referenced grants from co-investigators' FARs where the faculty
        appears as CoPI but their own FAR has stale status

    Per extraction rules:
      start_date in [Jan 1, Dec 31 of report_year]
      AND end_date >= Oct 1 of report_year (Q4 start)
      AND status = 'Funded - In Progress'
      AND role = PI or Co-PI
    """
    year_start = date(REPORT_YEAR, 1, 1)
    year_end   = date(REPORT_YEAR, 12, 31)
    canonical  = set()
    for g in far['grants']:
        status = g['status'].lower()
        if 'funded' not in status or 'progress' not in status:
            continue
        role = g['role']
        if role not in ('PI', 'CoPI'):
            continue
        start = parse_date(g['start'])
        if not start:
            continue
        # Grant must have started IN the report year
        if not (year_start <= start <= year_end):
            continue
        # End date, if present, must reach Q4 (grant still active in Oct–Dec)
        end = parse_date(g['end'])
        if end is not None and end < Q4_START:
            continue
        # Among qualifying grants, deduplicate by canonical title only
        # (same project may have multiple matching rows from continuation pages)
        canon = _canonical_title(g['title'])
        if canon not in canonical:
            canonical.add(canon)
            dbg(f"Grant: {g['title'][:55]} | {role} | {start}")

    far_count = len(canonical)

    # Add CV gifts (industry/foundation gifts not in the FAR grants table)
    gift_count = _parse_cv_gifts(cv_text)

    # Add cross-referenced grants from co-investigators' FARs
    xref_count = 0
    if faculty_last_name and all_far_data:
        xref_count = _cross_ref_grants(faculty_last_name, all_far_data)

    total = far_count + gift_count + xref_count
    dbg(f"Grants: far={far_count} gifts={gift_count} xref={xref_count} → {total}")
    return total


def _is_student_active(r):
    """
    Determine if a grad student is still active during the report year.
    Use End Semester column when available; fall back to Grad Date.
    Active = no end date, 'Ongoing', or end year > REPORT_YEAR.
    """
    end_sem = r.get('end_sem', '').strip()
    gd      = r.get('grad_date', '').strip()

    # End Semester is the most reliable signal
    if end_sem:
        if 'ongoing' in end_sem.lower():
            return True
        yr = semester_year(end_sem)
        if yr is not None:
            return yr > REPORT_YEAR

    # Fall back to grad_date
    if not gd or 'ongoing' in gd.lower():
        return True
    yr = semester_year(gd)
    if yr is not None:
        return yr > REPORT_YEAR
    # If graduation date exists but we can't parse it, assume inactive
    return False


def _cv_alumni_surnames(cv_text, max_year=REPORT_YEAR):
    """
    Extract surnames of graduated advisees from a CV.
    Returns surnames of anyone who graduated in year ≤ max_year.

    max_year=REPORT_YEAR  → used for Set A exclusion (includes report-year grads)
    max_year=REPORT_YEAR-1 → used for asterisk-path exclusion (keeps report-year grads
                              as "current" since they were active during the report year)

    Handles:
      A. Theses section numbered entries (Palermo-style multi-line)
      B. Defended inline list fused text (Qian-style)
      C. "Defended YEAR" with surrounding context
    """
    alumni = set()
    if not cv_text:
        return alumni

    # ── Format A: Theses section numbered entries ──
    theses_m = re.search(r'\nTheses?\b', cv_text, re.IGNORECASE)
    if theses_m:
        section_end = re.search(r'\n[A-Z][A-Z\s]{5,}\n', cv_text[theses_m.end():])
        theses_text = (cv_text[theses_m.end(): theses_m.end() + section_end.start()]
                       if section_end else cv_text[theses_m.end():])

        entry_re = re.compile(r'^\d+\.\s', re.MULTILINE)
        entry_starts = [m.start() for m in entry_re.finditer(theses_text)]
        for i, start in enumerate(entry_starts):
            end = entry_starts[i + 1] if i + 1 < len(entry_starts) else len(theses_text)
            entry = theses_text[start:end]
            yrs = re.findall(r'\b(20\d{2})\b', entry)
            if not any(int(y) <= max_year for y in yrs):
                continue
            nm = re.match(
                r'\d+\.\s*'
                r'(?:[A-Z][a-z\-]*(?:\.[A-Z\-][a-z\-]*)?\.\s+)*'
                r'([A-Z][a-z\-]+(?:\s+[A-Z][a-z\-]+)*)',
                entry.strip()
            )
            if nm:
                surname = nm.group(1).split()[-1].lower()
                if len(surname) > 2:
                    alumni.add(surname)

    # ── Format B: "Name(DefendedYEAR)" fused text ──
    for m in re.finditer(r'([A-Z][a-zA-Z\-]+)\s*\(?Defended\s*(\d{4})', cv_text):
        yr = int(m.group(2))
        if yr <= max_year:
            parts = re.findall(r'[A-Z][a-z\-]+', m.group(1))
            if parts:
                alumni.add(parts[-1].lower())

    # ── Format C: "Defended YEAR" with preceding name context ──
    _SKIP = {'phd', 'ph', 'ms', 'co', 'dr', 'prof', 'university', 'department',
             'tamu', 'usf', 'the', 'of', 'and', 'in', 'at', 'as', 'with'}
    for m in re.finditer(r'\bDefended\s+(\d{4})[,;.\)]', cv_text):
        yr = int(m.group(1))
        if yr <= max_year:
            before = cv_text[max(0, m.start() - 80):m.start()]
            words = re.findall(r'[A-Z][a-z\-]+', before)
            for word in reversed(words):
                wl = word.lower()
                if wl not in _SKIP and len(wl) > 2:
                    alumni.add(wl)
                    break

    return alumni


def rule6_ch_co(far, cv, cv_text=''):
    ga   = far['grad_advising']
    dept = faculty_home_dept(ga)
    home_pat = re.compile(r'\b' + re.escape(dept) + r'\b', re.IGNORECASE)

    # Pre-compute CV alumni surnames to exclude confirmed graduates from Set A.
    cv_alumni = _cv_alumni_surnames(cv_text)
    dbg(f"CH/CO CV alumni found: {cv_alumni}")

    # Set A: active chair/co-chair (Graduation_Date is null, 'Ongoing', or year > report_year)
    # When End Semester is available and graduation date is blank, use it as a proxy:
    # End Semester year > report_year → still active; End Semester year <= report_year → done.
    # Students with NO end_sem AND in CV alumni list are excluded (confirmed graduates).
    set_a = set()
    for r in ga:
        role = r['role'].strip()
        if not re.search(r'\b(chair|co.?chair)\b', role, re.IGNORECASE):
            continue
        gd      = r.get('grad_date', '').strip()
        end_sem = r.get('end_sem',   '').strip()
        # Primary check: graduation date
        if gd:
            if 'ongoing' in gd.lower():
                is_active = True
            else:
                yr = semester_year(gd)
                is_active = (yr is not None and yr > REPORT_YEAR)
        # Fallback: end semester (more reliable for Qian-style tables)
        elif end_sem:
            if 'ongoing' in end_sem.lower():
                is_active = True
            else:
                yr = semester_year(end_sem)
                is_active = (yr is None or yr > REPORT_YEAR)
        else:
            is_active = True   # no date info → assume active
        if is_active:
            name = r['name']
            if not name: continue
            surname = name.lower().split()[-1]
            # Exclude if CV alumni confirms graduation ≤ REPORT_YEAR,
            # BUT only when NO CV current student shares the same surname
            # (prevents wrongly excluding active students like Wang when an
            # old Yijie Wang alumna also has surname 'wang').
            cv_current = cv.get('current_advisees', set())
            if surname in cv_alumni and surname not in cv_current:
                dbg(f"CH/CO A: skipping {name} (CV alumni)")
                continue
            set_a.add(surname)

    # Set B: CV current advisees not already in A
    # Restrict to real names: 3-20 chars, no digits, not a compound word
    _GARBAGE_NAMES = {'postdoctoralresearchers', 'postdoctoral', 'researchers',
                      'gradstudents', 'currentstudents', 'students', 'advisees'}
    set_b = set()
    for surname in cv.get('current_advisees', set()):
        sn = surname.lower()
        if (2 <= len(sn) <= 20
                and not re.search(r'\d', sn)
                and sn not in _GARBAGE_NAMES
                and sn not in set_a):
            set_b.add(sn)

    # Set C: MS/MEN home-dept chair-graduates in report year only.
    # Per extraction_rules: "row.Degree matches pattern from Rule 4 (home-dept only)"
    # i.e. ^(MS|MEN)\s+{dept}$. PhD graduates this year are NOT in C (they would have
    # been in A while active; their graduation year = report_year falls through both sets).
    # Apply CV cross-reference to filter false chair assignments (e.g. co-supervised).
    ms_pat = re.compile(r'^(MS|MEN)\s+' + re.escape(dept) + r'$', re.IGNORECASE)
    set_c = set()
    for r in ga:
        role = r['role'].strip()
        if not re.search(r'\b(chair|co.?chair)\b', role, re.IGNORECASE):
            continue
        gd  = r.get('grad_date', '').strip()
        deg = r['degree'].upper()
        maj = r['major'].upper()
        combined = f"{deg} {maj}".strip()
        if date_contains_year(gd, REPORT_YEAR) and ms_pat.match(combined):
            name = r['name']
            if not name: continue
            # CV cross-reference: skip graduates not mentioned in CV
            if cv_text and not _name_in_cv(name, cv_text):
                dbg(f"CH/CO C: skipping {name} (not in CV)")
                continue
            set_c.add(name.lower().split()[-1])

    union = set_a | set_b | set_c
    dbg(f"CH/CO A={len(set_a)} B={len(set_b)} C={len(set_c)} → {len(union)}")
    dbg(f"  A={set_a}")
    dbg(f"  B={set_b}")
    return len(union)


def rule7_cp(far, cv, xlsx):
    far_count  = sum(1 for p in far['conf_pubs'] if p['year'] == REPORT_YEAR)
    cv_count   = cv.get('conf_pubs_2024', 0)
    xlsx_total = xlsx.get('conf_total', 0)
    result     = max(far_count, cv_count, xlsx_total)
    dbg(f"CP: far={far_count} cv={cv_count} xlsx={xlsx_total} → {result}")
    return result


def rule8_journals(xlsx, cv):
    xlsx_val = xlsx.get('journal_raw', 0)
    cv_count = cv.get('journal_pubs_2024', 0)
    result   = max(xlsx_val, cv_count)
    dbg(f"Journal: xlsx={xlsx_val} cv={cv_count} → {result}")
    return result


# ── File discovery ─────────────────────────────────────────────────────────────

def find_files(last_name, input_dir='.'):
    files = {}
    pattern = os.path.join(input_dir, f"F180Vita_*.{last_name}.pdf")
    matches = glob.glob(pattern)
    if not matches:
        matches = [f for f in glob.glob(os.path.join(input_dir, "F180Vita_*.pdf"))
                   if last_name.lower() in f.lower()]
    files['far'] = matches[0] if matches else None

    cv = os.path.join(input_dir, f"{last_name} CV.pdf")
    files['cv'] = cv if os.path.exists(cv) else None

    xl = os.path.join(input_dir, f"{last_name}.xlsx")
    files['xlsx'] = xl if os.path.exists(xl) else None

    staff = os.path.join(input_dir, "Faculty Support Staff (PostDoc and GARs).xlsx")
    files['staff'] = staff if os.path.exists(staff) else None

    return files


# ── Per-faculty extraction ─────────────────────────────────────────────────────

def extract_faculty(last_name, input_dir='.', api_key=None, all_far_data=None):
    files = find_files(last_name, input_dir)
    print(f"\n{'='*60}")
    print(f"Processing: {last_name}")
    for k, v in files.items():
        print(f"  {k}: {v or 'NOT FOUND'}")

    if not files['far']:
        print(f"  ERROR: FAR PDF not found for {last_name}")
        return None

    far      = parse_far(files['far'])
    cv       = parse_cv(files['cv']) if files['cv'] else {'current_advisees': set(), 'journal_pubs_2024': 0, 'conf_pubs_2024': 0}
    xlsx     = read_supplemental(files['xlsx']) if files['xlsx'] else {'ms_grad': 0, 'phd_grad': 0, 'journal_raw': 0, 'conf_total': 0}
    cv_text  = pdf_full_text(files['cv']) if files['cv'] else ''
    postdocs, gars = read_support_staff(files['staff'], last_name) if files['staff'] else (0, 0)

    ug   = rule1_ug_courses(far)
    grad = rule2_grad_courses(far)
    phd  = rule3_phd_graduated(far, xlsx, cv_text)
    ms   = rule4_ms_graduated(far, xlsx, cv_text)
    grts = rule5_grants(far, cv_text=cv_text,
                        faculty_last_name=last_name,
                        all_far_data=all_far_data)
    chco = rule6_ch_co(far, cv, cv_text)
    cp   = rule7_cp(far, cv, xlsx)
    jrnl = rule8_journals(xlsx, cv)

    result = {
        'last_name':  far.get('last_name', last_name),
        'first_name': far.get('first_name', ''),
        'title':      far.get('title', 'Professor'),
        'ug': ug, 'grad': grad, 'ms': ms, 'phd': phd,
        'grants': grts, 'ch_co': chco, 'cp': cp, 'journal': jrnl,
        'postdocs': postdocs, 'gars': gars,
    }
    print(f"\n  Results:")
    print(f"    UG={ug}  Grad={grad}  MS={ms}  PhD={phd}")
    print(f"    Grants={grts}  CH/CO={chco}  CP={cp}  Journal={jrnl}")
    return result


# ── Excel output ───────────────────────────────────────────────────────────────

def _border(color="D0D7E3"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg="2F5496"):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("FFFFFF")

def _dat(cell, even=True):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color="EEF2F8" if even else "FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def generate_excel(results, output_path):
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Summary sheet
    ws3 = wb.create_sheet("Summary", 0)
    ws3.sheet_properties.tabColor = "FF0000"
    ws3.merge_cells("A1:J1")
    c = ws3["A1"]
    c.value = f"FAR Extraction Results — CY{REPORT_YEAR}"
    c.font = Font(name="Arial", bold=True, size=14, color="2F5496")
    c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 28

    sum_hdrs = ["Last Name","First Name","UG","Grad","MS","PhD","Grants","CH/CO","CP","Journal"]
    ws3.row_dimensions[2].height = 20
    for c_idx, h in enumerate(sum_hdrs, 1):
        cell = ws3.cell(row=2, column=c_idx, value=h)
        _hdr(cell)
        ws3.column_dimensions[get_column_letter(c_idx)].width = 10
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 14

    for i, r in enumerate(results, 3):
        vals = [r['last_name'], r['first_name'], r['ug'], r['grad'],
                r['ms'], r['phd'], r['grants'], r['ch_co'], r['cp'], r['journal']]
        even = i % 2 == 0
        for c_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c_idx, value=v)
            _dat(cell, even)
    ws3.freeze_panes = "A3"

    # Teaching & Advising
    ws1 = wb.create_sheet("Teaching & Advising")
    ws1.sheet_properties.tabColor = "4472C4"
    h1 = ["Faculty Last Name","Faculty First Name","Title",
          "# UG\nCourses","# Grad\nCourses","# MS/MEN\nGraduated","# PhD\nGraduated"]
    ws1.row_dimensions[1].height = 40
    for c, (h, w) in enumerate(zip(h1, [18,16,14,10,10,12,12]), 1):
        cell = ws1.cell(row=1, column=c, value=h)
        _hdr(cell)
        ws1.column_dimensions[get_column_letter(c)].width = w
    for i, r in enumerate(results, 2):
        vals = [r['last_name'], r['first_name'], r['title'],
                r['ug'], r['grad'], r['ms'], r['phd']]
        even = i % 2 == 0
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=c, value=v)
            _dat(cell, even)
            if c <= 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.freeze_panes = "A2"

    # Research & Publications
    ws2 = wb.create_sheet("Research & Publications")
    ws2.sheet_properties.tabColor = "70AD47"
    h2 = ["Faculty Last Name","Faculty First Name",
          "Total\nGrants","CH/CO","CP\nTotals","Refereed\nJournal\nPapers"]
    ws2.row_dimensions[1].height = 40
    for c, (h, w) in enumerate(zip(h2, [18,16,10,10,10,12]), 1):
        cell = ws2.cell(row=1, column=c, value=h)
        _hdr(cell)
        ws2.column_dimensions[get_column_letter(c)].width = w
    for i, r in enumerate(results, 2):
        vals = [r['last_name'], r['first_name'],
                r['grants'], r['ch_co'], r['cp'], r['journal']]
        even = i % 2 == 0
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            _dat(cell, even)
            if c <= 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nSaved: {output_path}")



if __name__ == "__main__":
    main()
