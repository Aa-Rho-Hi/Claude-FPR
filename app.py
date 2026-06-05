import streamlit as st
import tempfile, os, glob, re, sys, io, json
from datetime import date

st.set_page_config(
    page_title="FAR Extraction Pipeline",
    page_icon="📄",
    layout="wide",
)

# ── Import extraction engine ───────────────────────────────────────────────────
try:
    import run_rules as rr
except Exception as e:
    st.error(f"❌ Failed to load run_rules.py: {e}")
    st.stop()

import counting  # deterministic, auditable entry-counting engine

# Map this app's long rule_type labels onto the engine's short filter codes.
_RULE_CODE = {
    "Count all entries in section":         "all",
    "Count entries containing keyword":     "contains",
    "Count entries from a specific year":   "year",
    "Count entries after year":             "after_year",
    "Count entries before year":            "before_year",
    "Count entries matching ALL keywords":  "all_of",
    "Count entries matching ANY keyword":   "any_of",
    "Count entries NOT containing keyword": "excludes",
}

# ══════════════════════════════════════════════════════════════════════════════
# Helper functions
# ══════════════════════════════════════════════════════════════════════════════

def _run_custom_rule(cv_text, cr):
    """Count entries for a standard (deterministic) custom rule.

    Delegates to the counting engine, which keeps segmentation, filtering, and
    counting separate: the section is sliced with robust header detection, split
    into discrete entries by a detector cascade (numbered / bracketed / bullets /
    blank-line blocks / flagged fallback) that merges multi-line entries, and the
    count is always len(matched). This fixes the multi-line-wrap over-count and
    the section-bleed over-count that line-based counting produced.

    Same signature and int return as before, so the crawl-all-PDFs / max-count
    loop keeps working unchanged.
    """
    rule = {
        "section":   cr.get("section", ""),
        "mode":      "regex",
        "rule_type": _RULE_CODE.get(cr.get("rule_type", ""), "contains"),
        "keywords":  [k.strip() for k in str(cr.get("keywords", "")).split(",") if k.strip()],
        "year":      str(cr.get("year", "")).strip(),
    }
    return counting.run_rule(cv_text, rule).count


def _strip_comment(line):
    """Remove a single leading comment marker (#, //, *) so a user who copies a
    commented example without deleting the # still gets a working rule."""
    return re.sub(r'^\s*(?:#+|//|\*)\s?', '', line)


# Anchors a rule block. Accepts "Rule name:", "Rule:", "Name:".
_RULE_ANCHOR = re.compile(r'(?im)^\s*(?:rule\s+name|rule|name)\s*:')


def _parse_rule_blocks(text):
    """Split comment-free `text` into (name, section, count) tuples."""
    out = []
    for block in _RULE_ANCHOR.split(text)[1:]:
        lines = [l.strip() for l in block.strip().split("\n") if l.strip()]
        if not lines:
            continue
        name, section, count = lines[0].strip(), "", ""
        for line in lines[1:]:
            ll = line.lower()
            if ll.startswith("look in:") or ll.startswith("section:") or ll.startswith("in:"):
                section = line.split(":", 1)[1].strip()
            elif ll.startswith("count:") or ll.startswith("counts:"):
                count = line.split(":", 1)[1].strip()
        if not name or not count:
            continue
        if name.startswith("[") or count.startswith("["):   # skip help-text template
            continue
        out.append((name, section, count))
    return out


def _parse_custom_rules(text):
    """
    Robust to real-world input: accepts "Rule name:"/"Rule:"/"Name:" and
    "Look in:"/"Section:"/"In:"; tolerates leftover comment markers (#, //);
    and if nothing is found after the 'ADD YOUR RULES HERE' marker, falls back
    to scanning the whole document (ignoring commented example lines).
    """
    rules = []

    marker = "ADD YOUR RULES HERE"
    idx = text.upper().find(marker.upper())
    region = text[idx + len(marker):] if idx >= 0 else text
    region_nc = "\n".join(_strip_comment(l) for l in region.split("\n"))
    parsed = _parse_rule_blocks(region_nc)

    if not parsed:   # user may have typed above the marker; scan all, drop comments
        no_comments = "\n".join(l for l in text.split("\n")
                                if not re.match(r'^\s*(?:#|//|\*)', l))
        parsed = _parse_rule_blocks(no_comments)

    for name, section, count in parsed:
        cl = count.lower()
        if cl.startswith("all entries") or cl == "all":
            rule_type, keywords, year = "Count all entries in section", "", ""
        elif cl.startswith("year:"):
            rule_type, keywords, year = "Count entries from a specific year", "", count.split(":",1)[1].strip()
        elif cl.startswith("any of:"):
            rule_type, keywords, year = "Count entries matching ANY keyword", count.split(":",1)[1].strip(), ""
        elif cl.startswith("all of:"):
            rule_type, keywords, year = "Count entries matching ALL keywords", count.split(":",1)[1].strip(), ""
        elif cl.startswith("excludes:") or cl.startswith("exclude:"):
            rule_type, keywords, year = "Count entries NOT containing keyword", count.split(":",1)[1].strip(), ""
        elif cl.startswith("contains:"):
            rule_type, keywords, year = "Count entries containing keyword", count.split(":",1)[1].strip(), ""
        else:
            # Try to satisfy common plain-English phrasings DETERMINISTICALLY (no
            # API key needed). Only genuinely semantic instructions fall through
            # to the AI per-entry classifier.
            m_after  = re.search(r'\b(?:after|since|from|>=?)\s*((?:19|20)\d{2})', cl)
            m_before = re.search(r'\b(?:before|prior to|earlier than|<=?)\s*((?:19|20)\d{2})', cl)
            m_year   = re.fullmatch(r'(?:in\s+|for\s+|year\s+)?((?:19|20)\d{2})', cl.strip())
            if m_before:
                rule_type, keywords, year = "Count entries before year", "", m_before.group(1)
            elif m_after:
                rule_type, keywords, year = "Count entries after year", "", m_after.group(1)
            elif m_year:
                rule_type, keywords, year = "Count entries from a specific year", "", m_year.group(1)
            elif len(count.split()) == 1:
                # A single bare word → keyword contains (deterministic, no AI).
                rule_type, keywords, year = "Count entries containing keyword", count, ""
            else:
                # Genuinely semantic → AI per-entry classifier (needs an API key).
                rule_type, keywords, year = "AI instruction", "", ""

        rules.append({"name": name, "section": section,
                      "rule_type": rule_type, "keywords": keywords, "year": year,
                      "count_raw": count})   # preserve original text for AI
    return rules


def _xlsx_to_summary(xlsx_path):
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(xlsx_path, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            lines.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(max_row=60, values_only=True):
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                if row_str.strip():
                    lines.append(row_str)
        return "\n".join(lines)
    except Exception:
        return ""


# ── AI extraction function ─────────────────────────────────────────────────────
_AI_SYSTEM_PROMPT = """\
Extract faculty metrics from annual report documents. Return ONLY a JSON object, no prose.

JSON keys required:
last_name, first_name, title (Professor/Associate Professor/Assistant Professor),
ug (int), grad (int), ms (int), phd (int), grants (int), ch_co (int), cp (int), journal (int)

Rules: count conservatively, integers only, 0 when absent, null for missing text fields.
Include any CUSTOM keys defined in the rules.
"""

def _call_ai(prompt, api_key, base_url, model):
    """
    Make a single AI call and return the text response.
    Handles both standard models (gpt-4o) and reasoning models (protected.gpt-5).
    """
    import requests as _req
    endpoint = (base_url.rstrip("/") + "/chat/completions"
                if base_url and base_url.strip()
                else "https://api.openai.com/v1/chat/completions")
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    # Detect reasoning model (o-series / protected.gpt-5)
    is_reasoning = any(x in model.lower() for x in ["o1", "o3", "o4", "gpt-5", "protected"])

    payload = {"model": model,
               "messages": [{"role": "user", "content": prompt}],
               "stream": False}
    if is_reasoning:
        payload["max_completion_tokens"] = 8000
    else:
        payload["max_tokens"] = 500
        payload["temperature"] = 0

    resp = _req.post(endpoint, json=payload, headers=headers, timeout=120)
    if not resp.ok:
        return None

    # Try standard JSON response first (gpt-4o, most models)
    try:
        data    = resp.json()
        content = (data.get("choices") or [{}])[0].get("message", {}).get("content", "")
        if content:
            return content.strip()
    except Exception:
        pass

    # Fall back to SSE parsing (some TAMU endpoints stream regardless)
    parts = []
    for raw in resp.text.splitlines():
        raw = raw.strip()
        if not raw: continue
        if raw.startswith("data:"):
            ps = raw[5:].strip()
            if ps == "[DONE]": continue
            try:
                ch  = json.loads(ps)
                c   = (ch.get("choices") or [{}])[0]
                msg = c.get("delta") or c.get("message") or {}
                parts.append(msg.get("content") or "")
            except Exception: pass
        elif raw.startswith("{"):
            try:
                obj = json.loads(raw)
                c   = (obj.get("choices") or [{}])[0]
                msg = c.get("message") or c.get("delta") or {}
                parts.append(msg.get("content") or "")
            except Exception: pass

    return "".join(parts).strip() or None


def _ai_classify_entries(entries, instruction, api_key, base_url, model, batch_size=15):
    """
    Ask the model to judge EACH entry match/no-match against `instruction`, then
    count the matches in Python. The model never returns a total — this is what
    eliminates the large-list miscount (13, then 36, then 55…), because nothing
    ever holds a running tally over a long list. Uses this app's own _call_ai so
    the configured endpoint / base_url / model (incl. reasoning models) still apply.
    Returns list[bool] aligned to `entries`, or raises on an unusable response.
    """
    flags = [False] * len(entries)
    for start in range(0, len(entries), batch_size):
        idxs = list(range(start, min(start + batch_size, len(entries))))
        listing = "\n".join(
            f"[{j}] " + re.sub(r'\s+', ' ', entries[g])[:500] for j, g in enumerate(idxs))
        prompt = (
            "You are labelling CV entries. For EACH entry below, decide whether it "
            f"satisfies this condition:\n\n  CONDITION: {instruction}\n\n"
            "Judge each entry independently and literally. Do NOT count or summarise. "
            "Return ONLY a JSON array, one object per entry, like "
            '[{"id":0,"match":true},{"id":1,"match":false}].\n\n'
            f"ENTRIES:\n{listing}")
        reply = _call_ai(prompt, api_key, base_url, model)
        if not reply:
            raise RuntimeError("empty AI reply")
        mt = re.search(r'\[.*\]', reply, re.DOTALL)
        if not mt:
            raise ValueError("no JSON array in AI reply")
        for item in json.loads(mt.group(0)):
            try:
                i = int(item.get("id"))
            except (TypeError, ValueError):
                continue
            if 0 <= i < len(idxs):
                flags[idxs[i]] = bool(item.get("match"))
    return flags


def extract_with_ai(rules_text, last_name, cv_text,
                    custom_rules, api_key, base_url=None, model="gpt-4o"):
    """
    Run each custom rule with a per-entry AI classifier. The section is segmented
    deterministically (counting.segment_entries), the model judges each entry,
    and PYTHON counts the matches. Returns {rule_name: count} (int) or
    {rule_name: None} on failure so the caller can fall back to the deterministic
    engine.
    """
    if not custom_rules:
        return {}

    results = {}
    for cr in custom_rules:
        name      = cr["name"]
        section   = cr.get("section", "")
        rule_type = cr.get("rule_type", "")
        keywords  = cr.get("keywords", "")
        year      = cr.get("year", "")

        # The condition each entry is judged against (plain-English when provided).
        raw = (cr.get("count_raw") or "").strip()
        if raw:
            instruction = raw
        elif "all entries" in rule_type.lower():
            instruction = "this is a distinct entry in the section"
        elif "year" in rule_type.lower() and year:
            instruction = f"the entry mentions the year {year}"
        elif "NOT" in rule_type:
            instruction = f"the entry does NOT contain '{keywords}'"
        elif "ANY" in rule_type:
            instruction = f"the entry contains ANY of: {keywords}"
        elif "ALL" in rule_type:
            instruction = f"the entry contains ALL of: {keywords}"
        else:
            instruction = f"the entry contains '{keywords}'"

        try:
            body, _w = counting.extract_section(cv_text, section)
            entries, _method, _conf = counting.segment_entries(body)
            if not entries:
                results[name] = 0
                continue
            flags = _ai_classify_entries(entries, instruction, api_key, base_url, model)
            results[name] = sum(1 for f in flags if f)   # Python counts the matches
        except Exception:
            results[name] = None  # signal fallback to the deterministic engine

    return results


# ── Default rules text ─────────────────────────────────────────────────────────
DEFAULT_RULES_TEXT = """\
Faculty Annual Report Extraction Rules
=======================================
Use these rules exactly and conservatively.
Standard rules always run. Scroll to the bottom to add your own custom rules.


Processing Workflow
--------------------
1. Upload all FAR PDFs, CV PDFs, and supplemental XLSX files together.
2. Select which faculty to process (all detected, or specific names).
3. Click Run Extraction. Each faculty member is processed in order.
4. Download the output Excel workbook when processing is complete.


Field Rules
------------
1. UG Courses: course numbers below 500, exclude research/seminar shells.
2. Grad Courses: course numbers 500 or above, same exclusions.
3. PhD Graduated: PhD degree, Chair role, graduation in report year.
4. MS/MSEN Graduated: MS/MSEN degree, Chair role, graduation in report year.
5. Total Grants: PI or CoPI, funded/active, start date in report year.
6. CH/CO: current doctoral advisees (Chair + Co-Chair).
7. CP Totals: conference papers in report year.
8. Refereed Journal Papers: peer-reviewed journals in report year.


CUSTOM RULES  (add your own below this line)
=============================================
Format:
   Rule name:  [column label]
   Look in:    [CV section — leave blank for whole document]
   Count:      all entries | contains: word | year: 2024 | any of: w1,w2 | excludes: word

Examples (remove # to activate):
   # Rule name:  Invited Talks
   # Look in:    Invited Talks
   # Count:      all entries

   # Rule name:  Book Chapters
   # Look in:    Book Chapters
   # Count:      all entries

ADD YOUR RULES HERE
====================
(write your custom rules below)

"""

# ══════════════════════════════════════════════════════════════════════════════
# Sidebar
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("⚙️ Settings")
    output_filename = st.text_input("Output filename", value="far_extraction_output.xlsx")

    st.markdown("---")
    st.markdown("#### 🤖 AI Extraction")
    st.caption(
        "Standard fields always use the rule-based pipeline. "
        "Provide an API key to power **custom rules** with AI."
    )
    ai_api_key  = st.text_input("AI API Key", type="password",
                                placeholder="sk-… or your TAMU key")
    ai_base_url = st.text_input("API Base URL (optional)",
                                placeholder="https://chat-api.tamu.ai/openai")
    ai_model    = st.text_input("Model name", value="protected.gpt-5")

    if ai_api_key:
        st.success("🟢 AI custom rules enabled")
    else:
        st.info("🔵 Rule-based extraction")

    st.markdown("---")
    st.markdown("**Expected file naming:**")
    st.code("F180Vita_F.Lastname.pdf  ← FAR\n"
            "Lastname CV.pdf          ← CV\n"
            "Lastname.xlsx            ← Supplemental\n"
            "support_staff.xlsx       ← Staff", language=None)

# ══════════════════════════════════════════════════════════════════════════════
# Main page
# ══════════════════════════════════════════════════════════════════════════════
st.title("📄 FAR Extraction Pipeline")

# ── 1. Upload ──────────────────────────────────────────────────────────────────
st.subheader("1. Upload Files")
uploaded_files = st.file_uploader(
    "Upload all files (FAR PDFs, CV PDFs, XLSX) — select all at once",
    accept_multiple_files=True,
    type=["pdf", "xlsx", "xls"],
)
if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} file(s) uploaded")
    with st.expander("Show uploaded files"):
        for f in uploaded_files:
            st.text(f"{f.name}  ({f.size/1024:.1f} KB)")

# ── 2. Faculty selection ───────────────────────────────────────────────────────
st.subheader("2. Select Faculty")
run_mode = st.radio("Process", ["All detected faculty", "Specific faculty only"], horizontal=True)
specific_names = ""
if run_mode == "Specific faculty only":
    specific_names = st.text_input("Last names (comma-separated)",
                                   placeholder="e.g. Narayanan, Palermo, Qian")

# ── 3. Configuration (collapsed) ──────────────────────────────────────────────
with st.expander("⚙️ Configuration", expanded=False):
    st.caption("Changes here apply to the next run.")
    cfg = st.session_state.get("cfg", {})

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Report Period**")
        cfg_year  = st.number_input("Report Year", min_value=2010, max_value=2100,
                                    value=cfg.get("report_year", rr.REPORT_YEAR), step=1)
        cfg_q4    = st.selectbox("Q4 starts in",  options=list(range(1,13)),
                                 index=cfg.get("q4_month",10)-1,
                                 format_func=lambda m: date(2000,m,1).strftime("%B"))
        cfg_gmin  = st.selectbox("Grant active through", options=list(range(1,13)),
                                 index=cfg.get("grant_min_month",10)-1,
                                 format_func=lambda m: date(2000,m,1).strftime("%B"))
    with col2:
        st.markdown("**Course Rules**")
        cfg_ugc   = st.number_input("UG course ceiling", min_value=100, max_value=900,
                                    value=cfg.get("ug_ceiling", rr.UG_COURSE_CEILING), step=100)
        cfg_shell = st.text_area("Exclude course title words",
                                 value=cfg.get("shell_tokens", ", ".join(sorted(rr.RESEARCH_SHELL_TOKENS))),
                                 height=80)

    st.markdown("**Grant Rules**")
    col3, col4 = st.columns(2)
    with col3:
        cfg_roles  = st.multiselect("Count grants where role is", ["PI","CoPI","Other"],
                                    default=cfg.get("grant_roles", sorted(rr.GRANT_COUNTED_ROLES)))
        cfg_gstatus = st.text_input("Grant status must contain",
                                    value=cfg.get("grant_status_kw", rr.GRANT_STATUS_KEYWORD))
    with col4:
        cfg_gprog = st.text_input("Grant status must also contain",
                                  value=cfg.get("grant_progress_kw", rr.GRANT_PROGRESS_KEYWORD))

    st.markdown("**Publication Headers**")
    col5, col6 = st.columns(2)
    with col5:
        cfg_jhdr = st.text_area("Journal headings",
                                value=cfg.get("journal_hdrs", ", ".join(sorted(rr.JOURNAL_HDR_KW))),
                                height=80)
    with col6:
        cfg_chdr = st.text_area("Conference headings",
                                value=cfg.get("conf_hdrs", ", ".join(sorted(rr.CONF_HDR_KW))),
                                height=80)

    if st.button("💾 Save Configuration", type="primary"):
        st.session_state["cfg"] = {
            "report_year": cfg_year, "q4_month": cfg_q4,
            "grant_min_month": cfg_gmin, "ug_ceiling": cfg_ugc,
            "shell_tokens": cfg_shell, "grant_roles": cfg_roles,
            "grant_status_kw": cfg_gstatus, "grant_progress_kw": cfg_gprog,
            "journal_hdrs": cfg_jhdr, "conf_hdrs": cfg_chdr,
        }
        st.success("✅ Configuration saved.")

# ── 4. Rules Editor (collapsed) ────────────────────────────────────────────────
with st.expander("📝 Rules Editor", expanded=False):
    st.caption("Add custom rules below the marker to create extra columns in the output.")
    rules_text = st.text_area("Rules", value=st.session_state.get("rules_text", DEFAULT_RULES_TEXT),
                              height=500, label_visibility="collapsed")
    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("💾 Save Rules", type="primary"):
            st.session_state["rules_text"]   = rules_text
            st.session_state["custom_rules"] = _parse_custom_rules(rules_text)
            n = len(st.session_state["custom_rules"])
            if n:
                st.success(f"✅ {n} custom rule{'s' if n!=1 else ''} saved.")
            else:
                st.warning(
                    "⚠️ Saved, but **0 custom rules** were recognized. Each rule needs a "
                    "`Rule name:` line and a `Count:` line, e.g.:\n\n"
                    "```\nRule name:  Patents Filed\nLook in:    Patents\nCount:      all entries\n```\n\n"
                    "If you copied an example, delete the leading `#`.")
    with c2:
        if st.button("↩️ Reset"):
            st.session_state["rules_text"]   = DEFAULT_RULES_TEXT
            st.session_state["custom_rules"] = []
            st.rerun()
    parsed = _parse_custom_rules(rules_text)
    if parsed:
        st.markdown(f"**{len(parsed)} custom rule(s) will add columns:**")
        for cr in parsed:
            st.markdown(f"- **{cr['name']}** — {cr['rule_type']} in *{cr['section'] or 'full doc'}*")

# ── 5. Run ─────────────────────────────────────────────────────────────────────
st.subheader("3. Run Extraction")
run_btn = st.button("▶ Run Extraction", type="primary", disabled=not uploaded_files)

if run_btn and uploaded_files:
    st.info(f"▶ Starting extraction — {len(uploaded_files)} file(s) received.")

    try:
        # Apply config overrides
        cfg = st.session_state.get("cfg", {})
        report_year = cfg.get("report_year", rr.REPORT_YEAR)
        rr.REPORT_YEAR            = report_year
        rr.Q4_START               = date(report_year, cfg.get("q4_month", 10), 1)
        rr.UG_COURSE_CEILING      = cfg.get("ug_ceiling", rr.UG_COURSE_CEILING)
        rr.GRANT_COUNTED_ROLES    = set(cfg.get("grant_roles", list(rr.GRANT_COUNTED_ROLES)))
        rr.GRANT_STATUS_KEYWORD   = cfg.get("grant_status_kw", rr.GRANT_STATUS_KEYWORD)
        rr.GRANT_PROGRESS_KEYWORD = cfg.get("grant_progress_kw", rr.GRANT_PROGRESS_KEYWORD)
        rr.GRANT_MIN_END_DATE     = date(report_year, cfg.get("grant_min_month", 10), 1)
        if cfg.get("shell_tokens"):
            rr.RESEARCH_SHELL_TOKENS = set(t.strip().upper() for t in cfg["shell_tokens"].split(",") if t.strip())
        if cfg.get("journal_hdrs"):
            rr.JOURNAL_HDR_KW = set(t.strip().lower() for t in cfg["journal_hdrs"].split(",") if t.strip())
        if cfg.get("conf_hdrs"):
            rr.CONF_HDR_KW = set(t.strip().lower() for t in cfg["conf_hdrs"].split(",") if t.strip())

        custom_rules = _parse_custom_rules(st.session_state.get("rules_text", DEFAULT_RULES_TEXT))

        tmpdir = tempfile.mkdtemp()
        st.write(f"📁 Temp directory created: `{tmpdir}`")

        # Write uploaded files
        saved = []
        for uf in uploaded_files:
            dest = os.path.join(tmpdir, uf.name)
            with open(dest, "wb") as fh:
                fh.write(uf.getbuffer())
            saved.append(uf.name)
        st.write(f"💾 Saved {len(saved)} files: {', '.join(saved)}")

        # Detect faculty
        far_files = glob.glob(os.path.join(tmpdir, "F180Vita_*.pdf"))
        st.write(f"🔍 Found {len(far_files)} FAR PDF(s): {[os.path.basename(f) for f in far_files]}")

        _known = getattr(rr, "KNOWN_FACULTY", ["Narayanan","Qian","Palermo","Hu","Duffield"])
        if run_mode == "All detected faculty":
            faculty_list = list(_known)
            for fp in far_files:
                m = re.match(r'F180Vita_\w+\.(\w+)\.pdf', os.path.basename(fp))
                if m and m.group(1) not in faculty_list:
                    faculty_list.append(m.group(1))
            faculty_list = [ln for ln in faculty_list
                            if glob.glob(os.path.join(tmpdir, f"F180Vita_*.{ln}.pdf"))]
        else:
            faculty_list = [n.strip() for n in specific_names.split(",") if n.strip()]

        st.write(f"👥 Faculty to process: {faculty_list}")

        if not faculty_list:
            st.error("❌ No faculty detected. FAR PDFs must be named `F180Vita_F.Lastname.pdf`.")
            st.stop()

        # Pre-parse FAR PDFs
        all_far_data = {}
        st.write("⏳ Pre-parsing FAR PDFs…")
        for far_path in glob.glob(os.path.join(tmpdir, "F180Vita_*.pdf")):
            m = re.match(r'F180Vita_\w+\.(\w+)\.pdf', os.path.basename(far_path))
            if not m: continue
            ln = m.group(1)
            try:
                all_far_data[ln] = (rr.parse_far(far_path), rr.pdf_full_text(far_path))
                st.write(f"  ✅ Parsed {os.path.basename(far_path)}")
            except Exception as e:
                st.warning(f"  ⚠️ Could not parse {os.path.basename(far_path)}: {e}")

        # Extract each faculty
        progress  = st.progress(0, text="Starting…")
        results   = []

        for i, last_name in enumerate(faculty_list):
            progress.progress(i / len(faculty_list), text=f"Processing {last_name}…")
            st.write(f"⏳ Processing **{last_name}**…")

            old_stdout, sys.stdout = sys.stdout, io.StringIO()
            try:
                r = rr.extract_faculty(last_name, input_dir=tmpdir,
                                       api_key=None, all_far_data=all_far_data)
            except Exception as e:
                r = None
                st.error(f"❌ {last_name} failed: {e}")
                import traceback; st.code(traceback.format_exc())
            finally:
                sys.stdout = old_stdout

            if r is None:
                continue

            # Custom rules — search ALL uploaded PDFs for this faculty member
            if custom_rules:
                all_texts = {}
                # FAR text
                far_txt = all_far_data.get(last_name, (None, ""))[1]
                if far_txt:
                    all_texts["FAR"] = far_txt
                # Every other PDF whose filename contains the last name
                for pdf_path in glob.glob(os.path.join(tmpdir, "*.pdf")):
                    fname = os.path.basename(pdf_path)
                    if last_name.lower() in fname.lower() and "F180Vita" not in fname:
                        try:
                            all_texts[fname] = rr.pdf_full_text(pdf_path)
                        except Exception:
                            pass

                std_types = {"Count all entries in section",
                             "Count entries containing keyword",
                             "Count entries from a specific year",
                             "Count entries after year",
                             "Count entries before year",
                             "Count entries matching ALL keywords",
                             "Count entries matching ANY keyword",
                             "Count entries NOT containing keyword"}

                for cr in custom_rules:
                    needs_ai = ai_api_key and cr.get("rule_type") not in std_types
                    if needs_ai:
                        best_text = max(all_texts.values(), key=len) if all_texts else ""
                        ai_res = extract_with_ai(
                            st.session_state.get("rules_text", DEFAULT_RULES_TEXT),
                            last_name, best_text,
                            [cr], ai_api_key,
                            ai_base_url or None, ai_model or "gpt-4o")
                        ai_val = (ai_res or {}).get(cr["name"])
                        r[cr["name"]] = ai_val if ai_val is not None else max(
                            (_run_custom_rule(txt, cr) for txt in all_texts.values()), default=0)
                    else:
                        # Run on every PDF and take the max
                        counts = [_run_custom_rule(txt, cr) for txt in all_texts.values()]
                        r[cr["name"]] = max(counts) if counts else 0

            results.append(r)
            st.write(f"✅ **{last_name}** — UG={r['ug']} Grad={r['grad']} "
                     f"MS={r['ms']} PhD={r['phd']} | "
                     f"Grants={r['grants']} CH/CO={r['ch_co']} "
                     f"CP={r['cp']} Journal={r['journal']}")

        progress.progress(1.0, text="Done!")

        if not results:
            st.error("❌ No results produced. Check file naming.")
        else:
            # Build Excel in memory using pandas + openpyxl
            try:
                import pandas as pd
                import io as _io

                # AR Memo Data sheet
                ar_rows = []
                for r in results:
                    first_initial = (r.get("first_name") or "X")[0]
                    row = {
                        "pdf_file_name":                 f"F180Vita_{first_initial}.{r['last_name']}.pdf",
                        "faculty_last_name":              r["last_name"],
                        "faculty_first_name":             r["first_name"],
                        "titles":                         r.get("title", ""),
                        "ug_courses_count":               r["ug"],
                        "grad_courses_count":             r["grad"],
                        "phd_students_graduated":         r["phd"],
                        "ms_msen_students_graduated":     r["ms"],
                        "total_grants":                   r["grants"],
                        "ch_co":                          r["ch_co"],
                        "cp_totals":                      r["cp"],
                        "refereed_journal_papers":        r["journal"],
                        "book_chapter":                   r.get("book_chapter", 0),
                        "textbook":                       r.get("textbook", 0),
                        "edited_book":                    r.get("edited_book", 0),
                        "number_of_postdocs":             r.get("postdocs", 0),
                        "number_of_graduate_assistants":  r.get("gars", 0),
                    }
                    for cr in custom_rules:
                        row[cr["name"]] = r.get(cr["name"], 0)
                    ar_rows.append(row)

                grants_rows = [{"Last name": r["last_name"], "First name": r["first_name"],
                                "Title": r.get("title",""), "External grants": r["grants"],
                                "Internal grants": 0, "Total grants": r["grants"]}
                               for r in results]

                service_rows = [{"pdf_file_name": f"F180Vita_{(r.get('first_name') or 'X')[0]}.{r['last_name']}.pdf",
                                 "faculty_last_name": r["last_name"],
                                 "faculty_first_name": r["first_name"],
                                 "admin_duties": ""}
                                for r in results]

                submission_rows = [{"Last name": r["last_name"], "First name": r["first_name"],
                                    "CV": 1, "FPR": 1} for r in results]

                buf = _io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    pd.DataFrame(ar_rows).to_excel(writer, sheet_name="AR Memo Data",   index=False)
                    pd.DataFrame(grants_rows).to_excel(writer, sheet_name="Grants",     index=False)
                    pd.DataFrame(service_rows).to_excel(writer, sheet_name="Service",   index=False)
                    pd.DataFrame(submission_rows).to_excel(writer, sheet_name="Submission", index=False)

                excel_bytes = buf.getvalue()
                st.session_state["results"]        = results
                st.session_state["excel_bytes"]    = excel_bytes
                st.session_state["excel_filename"] = output_filename
                st.session_state["custom_rules"]   = custom_rules
                st.success(f"✅ Excel ready — {len(excel_bytes):,} bytes across 4 sheets.")
            except Exception as e:
                import traceback
                st.error(f"❌ Excel generation failed: {e}")
                st.code(traceback.format_exc())

        # Cleanup tmpdir
        import shutil
        try: shutil.rmtree(tmpdir)
        except Exception: pass

    except Exception as e:
        import traceback
        st.error(f"❌ Unexpected error: {e}")
        st.code(traceback.format_exc())

# ── 6. Results (shown from session state — survives reruns) ───────────────────
if st.session_state.get("excel_bytes"):
    import pandas as pd
    _results      = st.session_state["results"]
    _custom_rules = st.session_state.get("custom_rules", [])
    _excel_bytes  = st.session_state["excel_bytes"]
    _fname        = st.session_state.get("excel_filename", "far_extraction_output.xlsx")

    st.markdown("---")
    st.subheader("📊 Results")

    rows = []
    for r in _results:
        row = {"Last Name": r["last_name"], "First Name": r["first_name"],
               "Title": r["title"], "UG": r["ug"], "Grad": r["grad"],
               "MS": r["ms"], "PhD": r["phd"], "Grants": r["grants"],
               "CH/CO": r["ch_co"], "CP": r["cp"], "Journal": r["journal"]}
        for cr in _custom_rules:
            row[cr["name"]] = r.get(cr["name"], 0)
        rows.append(row)

    st.dataframe(pd.DataFrame(rows).set_index("Last Name"), use_container_width=True)

    st.download_button(
        label=f"⬇️ Download {_fname}",
        data=_excel_bytes,
        file_name=_fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel",
    )
